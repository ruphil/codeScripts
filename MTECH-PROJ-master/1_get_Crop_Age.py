from openpyxl import load_workbook
import numpy as np

basepath = 'D:\MASTERS-PROJECT\MTECH-PROJ-Prob_1_VATI_vs_Field_Moisture\\'
in_file = basepath + 'Soil_Moisture_Berambadi_1617.xlsx'
out_file = basepath + 'Soil_Moisture_Berambadi_1617_CropAge_out.xlsx'

wb = load_workbook(in_file)
sheet = wb['SM_ML_values']

Crop_Type = list(sheet.columns)[5]

crop_array_count = 0
instance = 0
appending = 0
crops_array = []
previous_crop = ''
crop_array_temp = []
for crop in Crop_Type:
    row_num = crop.row
    if(row_num > 1 and crop.value is not None):
        # print (crop.value)
        current_crop = crop.value
        if (crop.value.strip() != 'No_Crop'):
            if(previous_crop != current_crop):
                if(previous_crop != 'No_Crop'):
                    crops_array.append(crop_array_temp)
                    crop_array_temp = []
                instance = 0
                appending = 1
            instance += 1
            crop_array_temp.append([crop.row, instance])
        # print (row_num, instance)
        if ((appending and crop.value.strip() == 'No_Crop')):
            crops_array.append(crop_array_temp)
            crop_array_temp = []
            appending = 0
            instance = 0
        
        if (row_num > 5265):
            break
        previous_crop = crop.value

for crop_set in crops_array:
    if (len(crop_set) != 0):
        max_instance = int(crop_set[len(crop_set)-1][1])
        for row in crop_set:
            sheet.cell(row=row[0], column=8).value = row[1]/max_instance
            
wb.save(out_file)
print ("done totally")